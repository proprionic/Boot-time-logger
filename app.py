import psutil
import datetime
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
from twilio.rest import Client  # Twilio API for sending WhatsApp messages

def boot_time():
    # -- Define timestamp and day --
    global today
    today = pd.Timestamp.today()  # Current date

    # day_of_week = today.day_name()
    global day_of_week
    day_of_week = today.day_name()[:3]

    # -- Save to txt --
    with open("boot_logs.txt", "a") as file:
        file.write(f"Boot time: {today} ({day_of_week})\n")

    # -- Save to Excel --
    filename = "boot_logs.xlsx"

    # Check if the file exists
    if os.path.exists(filename):
        workbook = load_workbook(filename)  # Load existing Excel file
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active

        # -- Set headers Monday - Sunday --
        headers = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        for i, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=i, value=header)
    
    # -- Get column for current day --
    day_columns = {'Mon': 1, 'Tue': 2, 'Wed': 3, "Thu": 4, 'Fri': 5, 'Sat': 6, 'Sun': 7}
    col = day_columns[day_of_week]

    # -- Find next empty row in column --
    next_row = sheet.max_row + 1

    # -- Append the date and time in the right column --
    sheet.cell(row=next_row, column=col, value=str(today.date()))  # Date
    sheet.cell(row=next_row + 1, column=col, value=str(today.time()))

    # -- Save file --
    workbook.save(filename)
    print(f"✅ Boot logged at: {today}")

    # Send WhatsApp message
    send_whatsapp_message(today, day_of_week)

def send_whatsapp_message(boot_time, day_of_week):
    # Twilio configuration (replace with your own values)
    account_sid = 'insert_accound_sid'
    auth_token = 'insert_auth_token'
    from_whatsapp = 'whatsapp:+'  # This is Twilio sandbox WhatsApp number
    to_whatsapp = 'whatsapp:+'  # Replace with your phone number

    client = Client(account_sid, auth_token)

    # Send the WhatsApp message
    message = client.messages.create(
        body=f"Boot time logged: {boot_time} ({day_of_week})",
        from_=from_whatsapp,
        to=to_whatsapp
    )

    print("✅ WhatsApp message sent.")

print("Logging boot time")
boot_time()
print(f"Boot time logged. {today} ({day_of_week})")
